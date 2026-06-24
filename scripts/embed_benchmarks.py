#!/usr/bin/env python3
"""Rebuild Compare against benchmarks from embedded dashboard DATA.

The HTML file already contains all journal metrics in ``const DATA``.
You do **not** need the xlsx to run this script.

What this script does:
  1. Computes benchmark series (all journals, client orgs, disciplines) from DATA
  2. Tags journals with top-level disciplines from ``data/ext_list*.csv`` (cols V–AZ)
  3. Optionally overlays corpus RTI/paper counts from xlsx ``by_year`` if present

Typical usage (CSV with discipline columns in data/):

    python3 scripts/embed_benchmarks.py
"""
from __future__ import annotations

import csv
import json
import re
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
SCRIPTS = Path(__file__).resolve().parent
if str(SCRIPTS) not in sys.path:
    sys.path.insert(0, str(SCRIPTS))
HTML = ROOT / "SciScore_journal_dashboard.html"
CLIENT_ORGS_PATH = ROOT / "scripts" / "client_orgs.json"
DEFAULT_XLSX = ROOT / "data" / "2026_sciscore_v3.xlsx"

# Top-level discipline flags in the xlsx (columns V–AZ).
DISCIPLINE_COL_START = 22  # V (1-based)
DISCIPLINE_COL_END = 52    # AZ inclusive (1-based)

RATE_KEYS = ["r", "sex", "pwr", "rand", "blind", "irb", "iacuc", "ab", "org", "cl", "tool", "data", "code", "prot", "data_id", "code_id"]
COUNT_KEYS = ["abn", "orgn", "cln", "tooln", "datan", "coden", "protn"]

BY_YEAR_ALIASES: dict[str, list[str]] = {
    "year": ["year", "pub_year", "publication year"],
    "n": ["n", "papers", "paper_count", "papers analysed", "papers analyzed", "count"],
    "r": ["r", "rti", "rigor", "rigor & transparency index", "rigor and transparency index"],
    "sex": ["sex", "sex_balance", "sex of subjects", "sex reporting"],
    "pwr": ["pwr", "power", "power analysis"],
    "rand": ["rand", "randomization", "randomisation"],
    "blind": ["blind", "blinding"],
    "irb": ["irb", "irb / ethics", "ethics", "irb approval"],
    "iacuc": ["iacuc"],
    "ab": ["ab", "antibodies", "antibody rrid", "antibodies w/ rrid"],
    "org": ["org", "organisms", "organism rrid", "organisms/models"],
    "cl": ["cl", "cell lines", "cell line rrid", "cell lines w/ rrid"],
    "tool": ["tool", "software", "software tools", "tools"],
    "abn": ["abn", "antibody detections"],
    "orgn": ["orgn", "organism detections"],
    "cln": ["cln", "cell line detections"],
    "tooln": ["tooln", "tool detections"],
}

JOURNAL_NAME_ALIASES = [
    "journal",
    "journal name",
    "journal title",
    "title",
    "name",
]


def find_xlsx(path: Path) -> Path | None:
    if path.exists():
        return path
    data_dir = ROOT / "data"
    if not data_dir.is_dir():
        return None
    matches = sorted(data_dir.glob("*.xlsx"))
    if len(matches) == 1:
        return matches[0]
    preferred = [p for p in matches if "sciscore" in p.name.lower() or "2026" in p.name]
    if len(preferred) == 1:
        return preferred[0]
    if matches:
        print("Multiple .xlsx files in data/ — using the first match:", matches[0], file=sys.stderr)
        return matches[0]
    return None


from html_json import extract_json_block, insert_or_replace_const_block, replace_const_block
from repair_dashboard_html import data_block_is_corrupt, repair_data_block, patch_init_hooks


def norm_header(value) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip().lower())


def looks_like_year(value) -> bool:
    if value is None or str(value).strip() == "":
        return False
    try:
        year = int(float(value))
    except (TypeError, ValueError):
        return False
    return 1990 <= year <= 2035


def infer_year_column(headers: list, data_rows: list) -> int | None:
    """by_year often uses a blank first-column header with years in the data rows."""
    if headers and norm_header(headers[0]) == "":
        hits = sum(1 for row in data_rows[:20] if row and looks_like_year(row[0]))
        if hits >= 2:
            return 0
    max_cols = max((len(row) for row in data_rows[:20]), default=len(headers))
    for idx in range(max_cols):
        hits = sum(
            1
            for row in data_rows[:20]
            if row and idx < len(row) and looks_like_year(row[idx])
        )
        if hits >= 2:
            return idx
    return None


def map_headers(headers: list, data_rows: list | None = None) -> dict[str, int]:
    normalized = [norm_header(h) for h in headers]
    mapping: dict[str, int] = {}
    for key, aliases in BY_YEAR_ALIASES.items():
        for alias in aliases:
            alias_norm = norm_header(alias)
            for idx, header in enumerate(normalized):
                if not header and key != "year":
                    continue
                if header == alias_norm or (header and alias_norm in header):
                    mapping[key] = idx
                    break
            if key in mapping:
                break
    if "year" not in mapping and data_rows:
        inferred = infer_year_column(headers, data_rows)
        if inferred is not None:
            mapping["year"] = inferred
    if "year" not in mapping:
        raise ValueError(f"Could not find a year column in by_year headers: {headers}")
    return mapping


def find_by_year_header_row(rows: list) -> tuple[int, list, dict[str, int]]:
    """Locate the header row when by_year has leading blank rows or multiple tables."""
    last_error: ValueError | None = None
    for idx in range(min(15, len(rows))):
        headers = [str(c) if c is not None else "" for c in rows[idx]]
        data_rows = rows[idx + 1 : idx + 21]
        if not data_rows:
            continue
        try:
            colmap = map_headers(headers, data_rows)
        except ValueError as exc:
            last_error = exc
            continue
        if "r" in colmap or "n" in colmap:
            return idx, headers, colmap
    if last_error:
        raise last_error
    headers = [str(c) if c is not None else "" for c in rows[0]]
    return 0, headers, map_headers(headers, rows[1:21])


def merge_by_year(primary: dict[str, dict], fallback: dict[str, dict]) -> dict[str, dict]:
    """Overlay sparse xlsx by_year rows onto computed corpus averages."""
    years = sorted(set(primary) | set(fallback))
    out: dict[str, dict] = {}
    for year in years:
        base = dict(fallback.get(year) or {})
        overlay = primary.get(year) or {}
        merged = {**base, **{k: v for k, v in overlay.items() if v is not None}}
        if (
            merged.get("r") is not None
            and merged["r"] < 1
            and (base.get("r") or 0) >= 1
        ):
            merged["r"] = base["r"]
        if merged:
            out[year] = merged
    return out


def parse_rate(value):
    if value is None or value == "":
        return None
    if isinstance(value, str):
        value = value.strip().replace("%", "")
        if not value:
            return None
    num = float(value)
    if num > 1.5:
        return num / 100.0
    return num


def parse_rti(value):
    """RTI is on a 0–10 scale; do not treat values like 4.2 as percentages."""
    if value is None or value == "":
        return None
    if isinstance(value, str):
        value = value.strip()
        if not value:
            return None
    return float(value)


def parse_int(value):
    if value is None or value == "":
        return None
    return int(float(value))


def row_to_year_data(row: list, colmap: dict[str, int]) -> dict:
    out: dict = {}
    for key in RATE_KEYS + COUNT_KEYS + ["n"]:
        if key not in colmap:
            continue
        raw = row[colmap[key]] if colmap[key] < len(row) else None
        if key in COUNT_KEYS or key == "n":
            out[key] = parse_int(raw)
        elif key == "r":
            out[key] = parse_rti(raw)
        else:
            out[key] = parse_rate(raw)
    return out


def read_by_year_xlsx(path: Path) -> dict[str, dict]:
    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError("openpyxl is required: pip install openpyxl") from exc

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if "by_year" not in wb.sheetnames:
        raise ValueError(f"Sheet 'by_year' not found in {path}; sheets: {wb.sheetnames}")
    ws = wb["by_year"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("by_year sheet is empty")

    header_idx, headers, colmap = find_by_year_header_row(rows)
    print("by_year columns mapped:", {k: headers[v] for k, v in sorted(colmap.items(), key=lambda x: x[1])})
    out: dict[str, dict] = {}
    for row in rows[header_idx + 1 :]:
        if not row or all(cell is None or str(cell).strip() == "" for cell in row):
            continue
        year_raw = row[colmap["year"]]
        if year_raw is None:
            continue
        year = str(int(float(year_raw))) if str(year_raw).replace(".", "", 1).isdigit() else str(year_raw).strip()
        out[year] = row_to_year_data(list(row), colmap)
    if not out:
        raise ValueError("No year rows parsed from by_year sheet")
    return out


def is_truthy_cell(value) -> bool:
    if value is None:
        return False
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value != 0
    text = str(value).strip().lower()
    if not text:
        return False
    return text in {"1", "true", "yes", "y", "x", "✓", "✔"} or text not in {"0", "false", "no", "n"}


def normalize_journal_name(value) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip())


def find_journal_list_csv() -> Path | None:
    data_dir = ROOT / "data"
    if not data_dir.is_dir():
        return None
    matches = sorted(
        {
            *data_dir.glob("ext_list*.csv"),
            *data_dir.glob("*ext_list*.csv"),
            *data_dir.glob("*May_2026*.csv"),
        }
    )
    if not matches:
        return None
    for path in matches:
        if "ext_list" in path.name.lower():
            return path
    return matches[0]


def discipline_columns(headers: list[str]) -> list[tuple[int, str]]:
    disc_start = DISCIPLINE_COL_START - 1
    disc_end = DISCIPLINE_COL_END
    cols: list[tuple[int, str]] = []
    for idx in range(disc_start, min(disc_end, len(headers))):
        label = str(headers[idx] or "").strip()
        if label:
            cols.append((idx, label))
    return cols


def extract_journal_disciplines(
    headers: list[str],
    rows: list[tuple],
    known_journals: set[str],
) -> dict[str, list[str]]:
    name_col = detect_journal_name_column(headers, rows, known_journals)
    if name_col is None:
        return {}

    discipline_headers = discipline_columns(headers)
    if not discipline_headers:
        return {}

    known_lower = {name.lower(): name for name in known_journals}
    out: dict[str, list[str]] = {}
    matched = 0
    for row in rows:
        if name_col >= len(row):
            continue
        raw_name = normalize_journal_name(row[name_col])
        if not raw_name:
            continue
        journal = raw_name if raw_name in known_journals else known_lower.get(raw_name.lower())
        if not journal:
            continue
        disciplines = []
        for idx, label in discipline_headers:
            if idx < len(row) and is_truthy_cell(row[idx]):
                disciplines.append(label)
        if disciplines:
            out[journal] = disciplines
            matched += 1
    return out


def discipline_order_from_headers(headers: list[str]) -> list[str]:
    return [label for _, label in discipline_columns(headers)]


def read_journal_disciplines_csv(
    path: Path, known_journals: set[str]
) -> tuple[dict[str, list[str]], list[str]]:
    with path.open(encoding="utf-8-sig", newline="") as handle:
        rows = [tuple(row) for row in csv.reader(handle)]
    if len(rows) < 2:
        return {}, []
    headers = [str(c).strip() if c is not None else "" for c in rows[0]]
    data_rows = rows[1:]
    result = extract_journal_disciplines(headers, data_rows, known_journals)
    order = discipline_order_from_headers(headers)
    print(
        f"Disciplines from {path.name}: {len(result)} journals tagged, "
        f"{len(order)} discipline columns (V–AZ)"
    )
    return result, order


def detect_journal_name_column(headers: list[str], rows: list[tuple], known_journals: set[str]) -> int | None:
    known_lower = {name.lower(): name for name in known_journals}
    best_idx = None
    best_score = 0
    search_cols = min(8, len(headers))
    for idx in range(search_cols):
        score = 0
        for row in rows[:500]:
            if idx >= len(row):
                continue
            name = normalize_journal_name(row[idx])
            if not name:
                continue
            if name in known_journals:
                score += 2
            elif name.lower() in known_lower:
                score += 2
        if score > best_score:
            best_score = score
            best_idx = idx
    return best_idx if best_score >= 5 else None


def read_journal_disciplines(path: Path, known_journals: set[str]) -> dict[str, list[str]]:
    """Read top-level discipline flags from columns V–AZ on the journal metadata sheet."""
    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError("openpyxl is required: pip install openpyxl") from exc

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    skip = {"by_year", "summary", "readme", "index"}
    best_sheet = None
    best_name_col = None
    best_rows: list[tuple] = []
    best_headers: list[str] = []

    for sheet_name in wb.sheetnames:
        if sheet_name.lower() in skip:
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            continue
        headers = [str(c).strip() if c is not None else "" for c in rows[0]]
        name_col = detect_journal_name_column(headers, rows[1:], known_journals)
        if name_col is None:
            continue
        score = sum(
            1
            for row in rows[1:501]
            if name_col < len(row)
            and normalize_journal_name(row[name_col]) in known_journals
        )
        if score > len(best_rows):
            best_sheet = sheet_name
            best_name_col = name_col
            best_rows = rows
            best_headers = headers

    if best_sheet is None or best_name_col is None:
        print("WARNING: could not locate a journal sheet with discipline columns V–AZ", file=sys.stderr)
        return {}, []

    result = extract_journal_disciplines(best_headers, best_rows[1:], known_journals)
    order = discipline_order_from_headers(best_headers)
    print(
        f"Disciplines from xlsx sheet '{best_sheet}': {len(result)} journals tagged, "
        f"{len(order)} discipline columns (V–AZ)"
    )
    return result, order


def aggregate_year(journals: dict, journal_names: list[str], year: str) -> dict | None:
    total_n = 0
    journal_count = 0
    weights = {k: 0.0 for k in RATE_KEYS}
    denominators = {k: 0.0 for k in RATE_KEYS}
    counts = {k: 0 for k in COUNT_KEYS}

    for name in journal_names:
        yd = journals.get(name, {}).get("y", {}).get(year)
        if not yd or not yd.get("n"):
            continue
        journal_count += 1
        total_n += int(yd["n"])
        for key in RATE_KEYS:
            if yd.get(key) is not None:
                weights[key] += float(yd[key]) * int(yd["n"])
                denominators[key] += int(yd["n"])
        for key in COUNT_KEYS:
            if yd.get(key) is not None:
                counts[key] += int(yd[key])

    if not total_n:
        return None
    out = {"n": total_n, "journalCount": journal_count}
    for key in RATE_KEYS:
        out[key] = weights[key] / denominators[key] if denominators[key] else None
    out.update(counts)
    return out


def benchmarks_for_journals(journals: dict, journal_names: list[str]) -> dict[str, dict]:
    years = sorted({year for name in journal_names for year in journals.get(name, {}).get("y", {})})
    out: dict[str, dict] = {}
    for year in years:
        agg = aggregate_year(journals, journal_names, year)
        if agg:
            out[year] = agg
    return out


def resolve_client_journals(journals: dict, client_cfg: dict) -> tuple[list[str], dict[str, list[str]]]:
    resolved: list[str] = []
    per_org: dict[str, list[str]] = {}
    all_names = set(journals.keys())

    for org, spec in client_cfg["orgs"].items():
        org_journals: list[str] = []
        for pub in spec.get("publishers", []):
            org_journals.extend(
                sorted(j for j, entry in journals.items() if entry.get("pub") == pub)
            )
        for journal in spec.get("journals", []):
            if journal in all_names:
                org_journals.append(journal)
        deduped = []
        seen = set()
        for journal in org_journals:
            if journal not in seen:
                seen.add(journal)
                deduped.append(journal)
        per_org[org] = deduped
        resolved.extend(deduped)

    final = []
    seen = set()
    for journal in resolved:
        if journal not in seen:
            seen.add(journal)
            final.append(journal)
    return final, per_org


def build_discipline_index(journals: dict, journal_disciplines: dict[str, list[str]]) -> dict[str, list[str]]:
    index: dict[str, list[str]] = {}
    for journal, disciplines in journal_disciplines.items():
        if journal not in journals:
            continue
        for discipline in disciplines:
            index.setdefault(discipline, []).append(journal)
    for discipline in index:
        index[discipline] = sorted(set(index[discipline]))
    return index


def apply_disciplines_to_data(journals: dict, journal_disciplines: dict[str, list[str]]) -> int:
    updated = 0
    for journal, disciplines in journal_disciplines.items():
        if journal not in journals:
            continue
        clean = sorted(set(d.strip() for d in disciplines if d and str(d).strip()))
        if clean:
            journals[journal]["disciplines"] = clean
            updated += 1
    return updated


def build_benchmark_catalog(
    client_cfg: dict,
    discipline_index: dict[str, list[str]],
    discipline_order: list[str],
) -> list[dict]:
    catalog: list[dict] = [
        {"id": "all", "label": "All journals", "group": "General"},
        {
            "id": "clients",
            "label": client_cfg.get("label", "SciScore clients"),
            "group": "General",
        },
    ]
    seen: set[str] = set()
    for discipline in discipline_order:
        if discipline_index.get(discipline):
            catalog.append(
                {
                    "id": f"discipline:{discipline}",
                    "label": discipline,
                    "group": "Disciplines",
                }
            )
            seen.add(discipline)
    for discipline in sorted(discipline_index):
        if discipline in seen or not discipline_index[discipline]:
            continue
        catalog.append(
            {
                "id": f"discipline:{discipline}",
                "label": discipline,
                "group": "Disciplines",
            }
        )
    return catalog


def build_benchmark_by_key(
    by_year: dict[str, dict],
    client_by_year: dict[str, dict],
    journals: dict,
    discipline_index: dict[str, list[str]],
) -> dict[str, dict[str, dict]]:
    out: dict[str, dict[str, dict]] = {"all": by_year, "clients": client_by_year}
    for discipline, names in discipline_index.items():
        if names:
            series = benchmarks_for_journals(journals, names)
            if series:
                out[f"discipline:{discipline}"] = series
    return out


def compute_by_year_fallback(journals: dict) -> dict[str, dict]:
    years = sorted({year for entry in journals.values() for year in entry.get("y", {})})
    out = {}
    for year in years:
        agg = aggregate_year(journals, list(journals.keys()), year)
        if agg:
            out[year] = agg
    return out


def main() -> None:
    requested = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_XLSX
    content = HTML.read_text(encoding="utf-8")
    if data_block_is_corrupt(content):
        print("Repairing corrupted const DATA block before embedding benchmarks…")
        content = repair_data_block(content)
    data = extract_json_block(content, "DATA")
    journals = data["j"]
    client_cfg = json.loads(CLIENT_ORGS_PATH.read_text(encoding="utf-8"))

    # All journal metrics already live in embedded DATA — xlsx is not required.
    by_year = compute_by_year_fallback(journals)
    source = "embedded_DATA"
    print("Using embedded DATA for all-journal benchmarks.")

    xlsx_path = find_xlsx(requested)
    if xlsx_path:
        try:
            by_year_xlsx = read_by_year_xlsx(xlsx_path)
            by_year = merge_by_year(by_year_xlsx, by_year)
            source = f"{xlsx_path.name}/by_year+embedded_DATA"
            print(f"Merged optional by_year overlay from {xlsx_path.name}")
        except Exception as exc:
            print(f"Note: skipped xlsx by_year ({exc}); using embedded DATA only")

    journal_disciplines: dict[str, list[str]] = {}
    discipline_order: list[str] = []
    csv_path = find_journal_list_csv()
    if csv_path:
        journal_disciplines, discipline_order = read_journal_disciplines_csv(
            csv_path, set(journals.keys())
        )
    elif xlsx_path:
        try:
            journal_disciplines, discipline_order = read_journal_disciplines(
                xlsx_path, set(journals.keys())
            )
        except Exception as exc:
            print(f"Note: skipped xlsx disciplines ({exc})")

    if journal_disciplines:
        tagged = apply_disciplines_to_data(journals, journal_disciplines)
        print(f"Applied discipline tags to {tagged} journals in embedded DATA")

    client_journal_names, per_org = resolve_client_journals(journals, client_cfg)
    client_by_year = benchmarks_for_journals(journals, client_journal_names)
    discipline_index = build_discipline_index(journals, journal_disciplines)
    benchmark_by_key = build_benchmark_by_key(
        by_year, client_by_year, journals, discipline_index
    )
    benchmark_catalog = build_benchmark_catalog(client_cfg, discipline_index, discipline_order)

    meta = {
        "source": source,
        "clientLabel": client_cfg.get("label", "SciScore clients"),
        "clientOrgs": list(client_cfg["orgs"].keys()),
        "clientJournalCount": len(client_journal_names),
        "clientJournalsByOrg": {org: len(items) for org, items in per_org.items()},
        "disciplineCount": len(discipline_index),
        "disciplinesTaggedJournalCount": len(journal_disciplines),
    }

    print("Client org journal counts:")
    for org, items in per_org.items():
        print(f"  {org}: {len(items)} journals")
    missing_orgs = [org for org, items in per_org.items() if not items]
    if missing_orgs:
        print(f"  (no journals matched: {', '.join(missing_orgs)})")
    if discipline_index:
        print(f"Discipline benchmarks: {len(discipline_index)} top-level fields")

    content = replace_const_block(content, "DATA", data)
    content = replace_const_block(content, "BY_YEAR_BENCHMARK", by_year)
    content = replace_const_block(content, "CLIENT_ORG_BENCHMARK", client_by_year)
    content = replace_const_block(content, "CLIENT_ORG_META", meta)
    content = replace_const_block(content, "CLIENT_ORG_JOURNALS", client_journal_names)
    content = insert_or_replace_const_block(
        content, "CLIENT_ORG_JOURNALS_BY_ORG", per_org, "CLIENT_ORG_JOURNALS"
    )
    content = insert_or_replace_const_block(
        content, "BENCHMARK_CATALOG", benchmark_catalog, "CLIENT_ORG_JOURNALS_BY_ORG"
    )
    content = insert_or_replace_const_block(
        content, "BENCHMARK_BY_KEY", benchmark_by_key, "BENCHMARK_CATALOG"
    )
    content = patch_init_hooks(content)
    HTML.write_text(content, encoding="utf-8")
    print(
        f"Embedded {len(benchmark_catalog)} compare options "
        f"and {len(benchmark_by_key)} benchmark series"
    )


if __name__ == "__main__":
    main()
