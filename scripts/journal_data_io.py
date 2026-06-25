"""Read 2026_sciscore_v3 journal-year metrics and build dashboard DATA."""
from __future__ import annotations

import csv
import sys
from collections import Counter, defaultdict
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
SCRIPTS = Path(__file__).resolve().parent
if str(SCRIPTS) not in sys.path:
    sys.path.insert(0, str(SCRIPTS))

from embed_benchmarks import find_xlsx  # noqa: E402

DEFAULT_CSV = ROOT / "data" / "2026_sciscore_v3.csv"
DEFAULT_XLSX = ROOT / "data" / "2026_sciscore_v3.xlsx"
EXT_LIST_GLOB = ("ext_list*.csv", "*ext_list*.csv", "*May_2026*.csv")

EMBEDDED_KEYS = {
    "n", "r", "sex", "pwr", "rand", "blind", "irb", "iacuc",
    "ab", "org", "cl", "tool", "abn", "orgn", "cln", "tooln",
}
NEW_KEYS = {"data", "code", "prot", "data_id", "code_id", "datan", "coden", "protn"}


def find_journal_csv() -> Path | None:
    data_dir = ROOT / "data"
    if not data_dir.is_dir():
        return None
    preferred = [
        data_dir / "2026_sciscore_v3.csv",
        data_dir / "by_journal_by_year.csv",
        *sorted(data_dir.glob("*sciscore*.csv")),
        *sorted(data_dir.glob("*by_journal*.csv")),
    ]
    seen: set[Path] = set()
    for path in preferred:
        if path in seen or not path.is_file():
            continue
        if _is_auxiliary_csv(path):
            continue
        seen.add(path)
        return path
    matches = sorted(
        p for p in data_dir.glob("*.csv")
        if p.is_file() and not _is_auxiliary_csv(p)
    )
    return matches[0] if matches else None


def _is_auxiliary_csv(path: Path) -> bool:
    name = path.name.lower()
    return "ext_list" in name or "by_year" in name


def find_journal_source(explicit: Path | None = None) -> Path | None:
    if explicit and explicit.exists():
        return explicit
    return find_journal_csv() or find_xlsx(DEFAULT_XLSX)


def find_ext_list_csv() -> Path | None:
    data_dir = ROOT / "data"
    if not data_dir.is_dir():
        return None
    for pattern in EXT_LIST_GLOB:
        matches = sorted(data_dir.glob(pattern))
        if matches:
            return matches[0]
    return None


def load_publisher_map() -> dict[str, str]:
    path = find_ext_list_csv()
    if not path:
        return {}
    mapping: dict[str, str] = {}
    with path.open(newline="", encoding="utf-8-sig") as handle:
        for row in csv.reader(handle):
            if not row or row[0] == "Sourcerecord ID":
                continue
            if len(row) < 2:
                continue
            title = row[1].strip()
            if not title:
                continue
            grouped = row[19].strip() if len(row) > 19 and row[19].strip() else ""
            publisher = grouped or (row[18].strip() if len(row) > 18 else "")
            if publisher:
                mapping[title] = publisher
    return mapping


def read_rows(path: Path) -> list[tuple]:
    if path.suffix.lower() == ".csv":
        with path.open(newline="", encoding="utf-8-sig") as handle:
            return [tuple(row) for row in csv.reader(handle)]
    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError("openpyxl is required for xlsx: pip install openpyxl") from exc
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = "by_journal_by_year" if "by_journal_by_year" in wb.sheetnames else wb.sheetnames[0]
    return list(wb[sheet].iter_rows(values_only=True))


def _num(value) -> float | None:
    if value is None:
        return None
    text = str(value).strip().replace("%", "").replace(",", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _int(value) -> int | None:
    num = _num(value)
    if num is None:
        return None
    return int(num)


def _rate(num: float | None, den: float | None) -> float | None:
    if num is None or den is None or den <= 0:
        return None
    return num / den


def _criterion_rate(row: list, idx: dict[str, int], prefix: str, n: int | None) -> float | None:
    detected = _num(row[idx[f"{prefix} detected"]])
    not_expected = _num(row[idx[f"{prefix} not detected expected"]])
    not_detected_not_expected = _num(row[idx[f"{prefix} not detected not expected"]])

    # In the v3 export, "* not detected expected" is often always zero.
    # Use pmid count as the denominator when available to avoid false 100% rates.
    if n is not None and n > 0:
        if detected is None and not_expected is None and not_detected_not_expected is None:
            return None
        return (detected or 0.0) / float(n)

    if detected is None and not_expected is None and not_detected_not_expected is None:
        return None
    detected = detected or 0.0
    denom = detected + (not_expected or 0.0) + (not_detected_not_expected or 0.0)
    if denom <= 0:
        return None
    return detected / denom


def _header_index(headers: list[str]) -> dict[str, int]:
    return {name: i for i, name in enumerate(headers)}


def is_sciscore_v3_export(headers: list[str]) -> bool:
    joined = {h.strip() for h in headers if h}
    return "pub year" in joined and "pmid count" in joined and "RTI" in joined


def row_to_metrics(row: list, idx: dict[str, int]) -> dict | None:
    n = _int(row[idx["pmid count"]])
    r = _num(row[idx["RTI"]])
    if not n and r is None:
        return None

    metrics: dict = {}
    if n is not None:
        metrics["n"] = n
    if r is not None:
        metrics["r"] = round(r, 1)

    for key, col in (
        ("sex", "% sex detected"),
        ("pwr", "% power detected"),
        ("rand", "% randomization detected"),
    ):
        val = _rate(_num(row[idx[col]]), 100.0) if _num(row[idx[col]]) is not None else None
        if val is not None:
            metrics[key] = val

    for key, prefix in (("blind", "blinding"), ("irb", "irb"), ("iacuc", "iacuc")):
        val = _criterion_rate(row, idx, prefix, n)
        if val is not None:
            metrics[key] = val

    resource_pairs = (
        ("ab", "antibody findable", "sum(antibody detected)", "abn"),
        ("org", "organism findable", "sum(organism detected)", "orgn"),
        ("cl", "cell line findable", "sum(cell line detected)", "cln"),
        ("tool", "tool findable", "sum(tool detected)", "tooln"),
    )
    for rate_key, findable_col, detected_col, count_key in resource_pairs:
        detected = _int(row[idx[detected_col]])
        findable = _int(row[idx[findable_col]])
        if detected is not None:
            metrics[count_key] = detected
        rate = _rate(float(findable or 0), float(detected or 0)) if detected else None
        if rate is not None:
            metrics[rate_key] = rate

    open_science = (
        ("data", "sum(data availability)", "datan"),
        ("code", "sum(code availability)", "coden"),
        ("prot", "sum(protocol id detected)", "protn"),
        ("data_id", "sum(data id detected)", None),
        ("code_id", "sum(code id detected)", None),
    )
    for rate_key, sum_col, count_key in open_science:
        total = _int(row[idx[sum_col]])
        if total is not None and count_key:
            metrics[count_key] = total
        if n and total is not None:
            metrics[rate_key] = total / n

    return metrics


def merge_year_metrics(existing: dict | None, incoming: dict) -> dict:
    if not existing:
        return incoming
    if int(incoming.get("n") or 0) >= int(existing.get("n") or 0):
        return incoming
    return existing


def build_data(rows: list[tuple]) -> tuple[dict, dict]:
    if not rows:
        raise ValueError("Input is empty")

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    if not is_sciscore_v3_export(headers):
        raise ValueError(f"Unrecognised journal export headers: {headers[:8]}")

    idx = _header_index(headers)
    journal_col = 0
    publishers = load_publisher_map()
    journals: dict[str, dict] = {}
    years = Counter()
    metric_presence: Counter[str] = Counter()

    for row in rows[1:]:
        if not row or not row[journal_col].strip():
            continue
        year_raw = row[idx["pub year"]].strip() if idx["pub year"] < len(row) else ""
        if not year_raw:
            continue
        year = str(int(float(year_raw))) if year_raw.replace(".", "", 1).isdigit() else year_raw
        metrics = row_to_metrics(list(row), idx)
        if not metrics:
            continue

        jname = row[journal_col].strip()
        pub = publishers.get(jname, "")
        entry = journals.setdefault(jname, {"pub": pub, "y": {}})
        if pub and not entry["pub"]:
            entry["pub"] = pub
        entry["y"][year] = merge_year_metrics(entry["y"].get(year), metrics)
        years[year] += 1
        for key, val in metrics.items():
            if val is not None:
                metric_presence[key] += 1

    publisher_index: dict[str, list[str]] = defaultdict(list)
    for jname, entry in journals.items():
        publisher_index[entry.get("pub") or "Unknown"].append(jname)

    meta = {
        "format": "2026_sciscore_v3.csv",
        "headers": headers,
        "journal_column": headers[journal_col],
        "years": dict(sorted(years.items(), key=lambda x: int(x[0]))),
        "metric_presence": dict(sorted(metric_presence.items())),
        "journal_count": len(journals),
        "row_count": sum(years.values()),
        "publisher_map_size": len(publishers),
    }
    return {
        "j": journals,
        "p": {pub: sorted(names) for pub, names in sorted(publisher_index.items())},
    }, meta


def inspect_rows(rows: list[tuple]) -> dict:
    _, meta = build_data(rows)
    sorted_years = sorted(meta["years"].keys(), key=int)
    present = set(meta["metric_presence"])
    samples = {}
    headers = rows[0]
    idx = _header_index([str(h).strip() if h else "" for h in headers])
    journal_col = 0
    for row in rows[1:]:
        if not row or not row[journal_col].strip():
            continue
        year = row[idx["pub year"]].strip()
        if year not in samples and year in sorted(meta["years"].keys()):
            metrics = row_to_metrics(list(row), idx)
            if metrics:
                samples[year] = metrics
        if len(samples) >= len(sorted_years):
            break

    return {
        "format": meta["format"],
        "row_count": meta["row_count"],
        "journal_count": meta["journal_count"],
        "years": {
            "min": sorted_years[0] if sorted_years else None,
            "max": sorted_years[-1] if sorted_years else None,
            "count": len(sorted_years),
            "rows_per_year": meta["years"],
        },
        "columns": {
            "headers": meta["headers"],
            "journal_column": meta["journal_column"],
        },
        "metrics": {
            "present_in_rows": meta["metric_presence"],
            "already_in_dashboard": sorted(EMBEDDED_KEYS),
            "new_vs_dashboard": sorted(present - EMBEDDED_KEYS),
            "missing_from_sheet": sorted(EMBEDDED_KEYS - present),
        },
        "samples": {
            "earliest_year": samples.get(sorted_years[0]) if sorted_years else None,
            "latest_year": samples.get(sorted_years[-1]) if sorted_years else None,
            "year_2015": samples.get("2015"),
            "year_2014": samples.get("2014"),
        },
    }
