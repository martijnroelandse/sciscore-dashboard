#!/usr/bin/env python3
"""Inspect 2026_sciscore_v3.xlsx sheets (especially by_journal_by_year).

Reports column headers, year coverage, metric availability, and columns not yet
mapped for dashboard embed. Run after placing the xlsx in data/:

    python3 scripts/inspect_xlsx.py
    python3 scripts/inspect_xlsx.py data/2026_sciscore_v3.xlsx
"""
from __future__ import annotations

import json
import sys
from collections import Counter
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
SCRIPTS = Path(__file__).resolve().parent
if str(SCRIPTS) not in sys.path:
    sys.path.insert(0, str(SCRIPTS))

from embed_benchmarks import (  # noqa: E402
    BY_YEAR_ALIASES,
    COUNT_KEYS,
    JOURNAL_NAME_ALIASES,
    RATE_KEYS,
    find_by_year_header_row,
    find_xlsx,
    map_headers,
    norm_header,
    parse_int,
    parse_rate,
    row_to_year_data,
)

DEFAULT_XLSX = ROOT / "data" / "2026_sciscore_v3.xlsx"

# Candidate aliases for newer SciScore v3 export columns (confirmed at inspect time).
EXTRA_ALIASES: dict[str, list[str]] = {
    "data": ["data", "data availability", "data_avail", "data availability statement"],
    "code": ["code", "code availability", "code_avail", "code availability statement"],
    "prot": ["prot", "protocol", "protocol availability", "protocol identifier", "protocol identifiers"],
    "data_id": ["data identifier", "data identifiers", "data_id", "data id"],
    "code_id": ["code identifier", "code identifiers", "code_id", "code id"],
    "prot_id": ["protocol id", "protocol ids", "prot_id"],
    "datan": ["data detections", "data_n", "data count"],
    "coden": ["code detections", "code_n", "code count"],
    "protn": ["protocol detections", "protocol_n", "protocol count"],
    "stats": ["stats", "statistics", "statistical reporting", "statistics module"],
    "publisher": ["publisher", "pub", "publisher name"],
}


def load_workbook(path: Path):
    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError("openpyxl is required: pip install openpyxl") from exc
    return openpyxl.load_workbook(path, read_only=True, data_only=True)


def map_all_headers(headers: list, data_rows: list) -> dict[str, int]:
    mapping = map_headers(headers, data_rows)
    normalized = [norm_header(h) for h in headers]
    for key, aliases in EXTRA_ALIASES.items():
        if key in mapping:
            continue
        for alias in aliases:
            alias_norm = norm_header(alias)
            for idx, header in enumerate(normalized):
                if header == alias_norm or (header and alias_norm in header):
                    mapping[key] = idx
                    break
            if key in mapping:
                break
    return mapping


def detect_journal_column(headers: list, data_rows: list) -> int | None:
    normalized = [norm_header(h) for h in headers]
    for alias in JOURNAL_NAME_ALIASES:
        alias_norm = norm_header(alias)
        for idx, header in enumerate(normalized):
            if header == alias_norm or (header and alias_norm in header):
                return idx
    # Heuristic: first text column with high cardinality in sample rows.
    for idx in range(min(8, len(headers))):
        values = [
            str(row[idx]).strip()
            for row in data_rows[:200]
            if row and idx < len(row) and row[idx] not in (None, "")
        ]
        if len(values) >= 20 and len(set(values)) >= 15:
            return idx
    return None


def parse_row_metrics(row: list, colmap: dict[str, int]) -> dict:
    out = row_to_year_data(list(row), colmap)
    for key in EXTRA_ALIASES:
        if key in colmap and key not in out:
            raw = row[colmap[key]] if colmap[key] < len(row) else None
            if key in ("datan", "coden", "protn"):
                out[key] = parse_int(raw)
            elif key == "publisher":
                out[key] = str(raw).strip() if raw not in (None, "") else None
            else:
                out[key] = parse_rate(raw)
    return out


def inspect_by_journal_by_year(ws) -> dict:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("by_journal_by_year sheet is empty")

    header_idx, headers, _ = find_by_year_header_row(rows)
    data_rows = rows[header_idx + 1 :]
    colmap = map_all_headers(headers, data_rows[:50])
    journal_col = detect_journal_column(headers, data_rows[:200])

    years = Counter()
    journals = set()
    publishers = Counter()
    metric_presence: dict[str, int] = Counter()
    unmapped_headers = [
        headers[i]
        for i, h in enumerate(headers)
        if h not in (None, "")
        and i not in colmap.values()
        and norm_header(h) not in {"", "unnamed"}
    ]

    sample_by_year: dict[str, dict] = {}

    for row in data_rows:
        if not row or all(cell is None or str(cell).strip() == "" for cell in row):
            continue
        if "year" not in colmap:
            continue
        year_raw = row[colmap["year"]]
        if year_raw is None:
            continue
        year = str(int(float(year_raw))) if str(year_raw).replace(".", "", 1).isdigit() else str(year_raw).strip()
        metrics = parse_row_metrics(list(row), colmap)
        if not metrics.get("n") and metrics.get("r") is None:
            continue

        years[year] += 1
        if journal_col is not None and journal_col < len(row) and row[journal_col]:
            journals.add(str(row[journal_col]).strip())
        pub = metrics.get("publisher")
        if not pub and "publisher" in colmap and colmap["publisher"] < len(row):
            pub = row[colmap["publisher"]]
        if pub:
            publishers[str(pub).strip()] += 1

        for key, val in metrics.items():
            if key != "publisher" and val is not None:
                metric_presence[key] += 1

        if year not in sample_by_year and metrics:
            sample_by_year[year] = metrics

    sorted_years = sorted(years.keys(), key=int)
    embedded_keys = set(RATE_KEYS + COUNT_KEYS + ["n"])

    return {
        "sheet": "by_journal_by_year",
        "header_row": header_idx + 1,
        "row_count": sum(years.values()),
        "journal_count": len(journals),
        "publisher_count": len(publishers),
        "years": {
            "min": sorted_years[0] if sorted_years else None,
            "max": sorted_years[-1] if sorted_years else None,
            "count": len(sorted_years),
            "rows_per_year": dict(sorted(years.items(), key=lambda x: int(x[0]))),
        },
        "columns": {
            "headers": [str(h) if h is not None else "" for h in headers],
            "mapped": {k: headers[v] for k, v in sorted(colmap.items(), key=lambda x: x[1])},
            "journal_column": headers[journal_col] if journal_col is not None else None,
            "unmapped": unmapped_headers,
        },
        "metrics": {
            "present_in_rows": dict(sorted(metric_presence.items())),
            "already_in_dashboard": sorted(embedded_keys),
            "new_vs_dashboard": sorted(set(metric_presence) - embedded_keys - {"publisher"}),
            "missing_from_sheet": sorted(embedded_keys - set(metric_presence)),
        },
        "samples": {
            "earliest_year": sample_by_year.get(sorted_years[0]) if sorted_years else None,
            "latest_year": sample_by_year.get(sorted_years[-1]) if sorted_years else None,
            "year_2015": sample_by_year.get("2015"),
            "year_2014": sample_by_year.get("2014"),
        },
    }


def inspect_by_year(ws) -> dict:
    rows = list(ws.iter_rows(values_only=True))
    header_idx, headers, colmap = find_by_year_header_row(rows)
    colmap = map_all_headers(headers, rows[header_idx + 1 : header_idx + 21])
    years = []
    for row in rows[header_idx + 1 :]:
        if not row or "year" not in colmap:
            continue
        year_raw = row[colmap["year"]]
        if year_raw is None:
            continue
        year = str(int(float(year_raw)))
        metrics = parse_row_metrics(list(row), colmap)
        years.append((year, sorted(metrics.keys())))
    return {
        "sheet": "by_year",
        "years": len(years),
        "year_range": f"{years[0][0]}–{years[-1][0]}" if years else None,
        "full_metrics_from": next((y for y, keys in years if len(keys) > 3), None),
        "columns_mapped": {k: headers[v] for k, v in sorted(colmap.items(), key=lambda x: x[1])},
    }


def main() -> int:
    path = Path(sys.argv[1]) if len(sys.argv) > 1 else find_xlsx(DEFAULT_XLSX)
    if not path or not path.exists():
        print("No xlsx found. Place 2026_sciscore_v3.xlsx in data/ and re-run.", file=sys.stderr)
        print(f"  Expected: {DEFAULT_XLSX}", file=sys.stderr)
        return 1

    wb = load_workbook(path)
    print(f"File: {path}")
    print(f"Sheets: {', '.join(wb.sheetnames)}\n")

    report: dict = {"file": str(path), "sheets": wb.sheetnames}

    if "by_journal_by_year" in wb.sheetnames:
        report["by_journal_by_year"] = inspect_by_journal_by_year(wb["by_journal_by_year"])
    else:
        print("WARNING: sheet 'by_journal_by_year' not found", file=sys.stderr)

    if "by_year" in wb.sheetnames:
        report["by_year"] = inspect_by_year(wb["by_year"])

    print(json.dumps(report, indent=2, default=str))

    # Summary block for humans
    bj = report.get("by_journal_by_year")
    if bj:
        y = bj["years"]
        print("\n--- Summary ---")
        print(f"Journals: {bj['journal_count']:,}  |  Rows: {bj['row_count']:,}")
        print(f"Year range: {y['min']}–{y['max']} ({y['count']} years)")
        new = bj["metrics"]["new_vs_dashboard"]
        if new:
            print(f"New metrics vs dashboard: {', '.join(new)}")
        if bj["columns"]["unmapped"]:
            print(f"Unmapped columns: {', '.join(str(c) for c in bj['columns']['unmapped'][:12])}")
            if len(bj["columns"]["unmapped"]) > 12:
                print(f"  … and {len(bj['columns']['unmapped']) - 12} more")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
